from openpyxl import *
import pandas as pd
import os
import re
from traits.api import *
from traitsui.api import *

class APS_Req_Updater(HasTraits):
    '''Below assay_folder and comp_assay_req_log are class level variable and every class will have the same values for these
    just need to remember in writing the methods that inside of their definition hjas to refer with self.<attribute>'''

    def __init__(self):
        self.assay_folder = {'GI':r'N:\Departments\R_D\NexGen\Assay Integration\04-APS File Requests\Gastro Intestinal Protocols\Requests',
                        'BCID-GP':r'N:\Departments\R_D\NexGen\Assay Integration\04-APS File Requests\BCID\BCID-GP\Requests',
                        'BCID-GN':r'N:\Departments\R_D\NexGen\Assay Integration\04-APS File Requests\BCID\BCID-GN\Verified Requests',
                        'BCID-FP':r'N:\Departments\R_D\NexGen\Assay Integration\04-APS File Requests\BCID\BCID-FP\Verified Requests',
                        'RP':r'N:\Departments\R_D\NexGen\Assay Integration\04-APS File Requests\RP\Requests',
                        'LRTI':r'N:\Departments\R_D\NexGen\Assay Integration\04-APS File Requests\LRTI\Requests'}
        self.comp_assay_req_log = {'GI':r'M:\Sandbox\GI\Protocols since Nov 2016.xlsx',
                        'BCID-GP':r'M:\Sandbox\BCID-GP\Request GP-000 BASELINE-GS-1.xlsx',
                        'BCID-GN':r'M:\Sandbox\BCID-GN\GN-0007_BCID-GN-ControlEffectiveness-v2-1-TC03.xlsx',
                        'BCID-FP':r'M:\Sandbox\BCID-FP\FP-009-FP-0026 opus.xlsx',
                        'RP':r'M:\Sandbox\RP\RP-Tracking-sheet-07102017.xlsx'}

        # {'GI':r'N:\Departments\R_D\NexGen\Assay Integration\04-APS File Requests\Gastro Intestinal Protocols\Protocols since Nov 2016.xlsx',
        #                 'BCID-GP':r'N:\Departments\R_D\NexGen\Assay Integration\04-APS File Requests\BCID\06012017_BCID APS OPUS TRACKING SHEET.xlsx',
        #                 'BCID-GN':r'N:\Departments\R_D\NexGen\Assay Integration\04-APS File Requests\BCID\06012017_BCID APS OPUS TRACKING SHEET.xlsx',
        #                 'BCID-FP':r'N:\Departments\R_D\NexGen\Assay Integration\04-APS File Requests\BCID\06012017_BCID APS OPUS TRACKING SHEET.xlsx',
        #                 'RP':r'N:\Departments\R_D\NexGen\Assay Integration\04-APS File Requests\RP\RP-Tracking-sheet-07102017.xlsx',
        #                 'LRTI':r'N:\Departments\R_D\NexGen\Assay Integration\04-APS File Requests\LRTI\APS_Requests.xlsx'}

    def update_assay_reqs(self, assay, field2='field'):

        update_protocols = list()
        aps_req_list = list()
        _temp_log = list()
        logged_opi = set()
        opus_dev_log = list()
        ### Read a list of all protocols in a given directory
            #returns list of everything in current working directory
        dir_list = os.listdir(self.assay_folder[assay])

        for req in dir_list:
           if assay in req:
                aps_req_list.append(req)
        if len(aps_req_list) > 1:
                print(aps_req_list[0])

        '''Reading from protocol log sheet to compare, what's listed in the
        Request directory vs. what's already been logged. If any requests present that've
        not been logged, record filenames to then take catalog impt info and write to
        log sheet '''

        # open Xl sheet to read in already logged requests.
        comp_req_wb = load_workbook(self.comp_assay_req_log[assay])

        print(comp_req_wb['Sheet1'].rows)
        for cell in tuple(comp_req_wb['Sheet1'].rows):
        #    print(cell[0].value)
            if cell[0].value is not None and str(assay) in cell[0].value:
                _temp_log.append(cell[0].value)
        #checks if the cell value isn't None and has 'GI' to only include only OPUS names w/ no trivial excel formatting

        '''IMPT detail here that however the assays are described/labeled here should be uniform
        the discrepancy between how GI and RP were labeled threw an error'''

        for entry in _temp_log:
            # diagnostic tool print(assay,entry)
            # try:
            opus_req = re.match("({}-\\d*).*".format(assay),entry)
            # diagnostic tool print(opus_req.groups(1))
            logged_opi.add(opus_req.groups(1)[0])
            # except:
            #     continue
        #diagnostic tool print(_temp_log[0], 'temp_log')
        #diagnostic tool print(logged_opi)
        #only return the set of unique values from the list

        '''Need now open iteratively files, from the dri_list, with the reqs_to_update
        that are not yet catalogged. Either join path or b/c already have list, go through
        and if the pulled regex NOT in logged_opi. open to take info have correct list write
        to the comp_assay_req_log'''

        req_bk = load_workbook(self.comp_assay_req_log[assay])
        req_sheet = req_bk.active
        # print(req_bk.active)
        # # for rows in req_sheet.rows:
        # #     for cell in rows:
        # #         print(cell.value)
        #
        # # Access specific values from
        # req_sheet['D28'] = str(aps_req_list[0])
        # req_bk.save(self.comp_assay_req_log[assay])


# ### Filter out protocols already updated
# reqs_to_update = list()
# for entry in aps_req_list:
#     #checks if no GI skips that iteration to prevent being flagged and time waste
#     if 'GI' not in entry or re.search(".*GI-\d\d\d\d.+",entry) is None: continue
#     #extracts matching seq as a list, to access raw string b/c = list need index the element
#     if re.findall(".*(GI-\d\d\d\d).+",entry)[0] not in logged_opi:
#         reqs_to_update.append(entry)
#
# ### Access files that've not yet been updated on the dev log sheet
# os.chdir(r'N:\Departments\R_D\NexGen\Assay Integration\04-APS File Requests\Gastro Intestinal Protocols\Requests')
# # creates the absolute file path for for the protocols to be updated
# fi_path = os.path.join(os.getcwd(),reqs_to_update[2])
#
# ###Open Request to update
#
# req_bk = load_workbook(filename = fi_path)
# req_sheet = req_bk.active
#
# for rows in req_sheet.rows:
#     for cell in rows:
#         print(cell.value)
#
# # Access specific values from
# req_sheet['C8'].value


APS_Req_Updater().update_assay_reqs('RP')
